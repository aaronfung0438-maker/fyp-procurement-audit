"""
Generate academic result tables (Excel, one sheet per table).
Output: data/final/results_tables.xlsx

Human data = PRIMARY (left columns, bold).
LLM simulation = SECONDARY reference (right columns, lighter).

Tables:
  T1  Stage 4 AI Baseline Performance
  T2  Accuracy Overview — Human first, LLM alongside (RQ1)
  T3  Effect Sizes — Human first, LLM alongside (RQ1)
  T4  Accuracy by Signal Type — Human first, LLM alongside (RQ2)
  T5  AI Override Rate — Human first, LLM alongside (RQ3)
  T6  Per-order Override Detail (RQ3)
"""

import json
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────────────────────
BASE    = Path("data")
HUMAN   = BASE / "final" / "FYP Procurement Audit Data.xlsx"
LLM_CSV = BASE / "llm_sim" / "llm_sim_results.csv"
KEY     = sorted(BASE.rglob("experiment_32qs_*_KEY.xlsx"))[-1]
G2V     = sorted((BASE / "stage4").glob("g2_verdicts_exp_*.json"))[-1]
OUT     = BASE / "final" / "results_tables.xlsx"

print(f"KEY file : {KEY.name}")
print(f"G2V file : {G2V.name}")

# ── load KEY ───────────────────────────────────────────────────────────────────
key_df = pd.read_excel(KEY, usecols=["po_id", "experiment_block", "injection_plan"])
key_df["truth"] = key_df["injection_plan"].apply(
    lambda x: "normal" if str(x).lower() == "none" else "anomaly")
truth_map = dict(zip(key_df["po_id"], key_df["truth"]))

print(f"\nKEY rows: {len(key_df)}, truth counts:\n{key_df['truth'].value_counts().to_string()}")

# ── load human data ────────────────────────────────────────────────────────────
sess  = pd.read_excel(HUMAN, sheet_name="sessions")
exp_h = sess[sess["phase"] == "experiment"].copy()
before_n = len(exp_h)
exp_h = exp_h.merge(key_df[["po_id", "truth", "experiment_block"]], on="po_id", how="left")

print(f"\nHuman experiment rows: {before_n} → after merge: {len(exp_h)}")
print(f"truth NaN after merge: {exp_h['truth'].isna().sum()}")
print(f"truth value counts:\n{exp_h['truth'].value_counts(dropna=False).to_string()}")
print(f"Sample po_ids (sessions): {exp_h['po_id'].head(3).tolist()}")
print(f"Sample po_ids (KEY):      {key_df['po_id'].head(3).tolist()}")

exp_h["pred"]    = exp_h["judgment"].astype(str).str.lower().str.strip()
exp_h["pred"]    = exp_h["pred"].apply(lambda j: "anomaly" if j == "suspicious" else j)
exp_h["correct"] = exp_h["pred"] == exp_h["truth"]
exp_h["is_sus"]  = exp_h["judgment"].astype(str).str.lower().str.strip() == "suspicious"

print(f"\nHuman correct rate overall: {exp_h['correct'].mean():.3f}")
print("Human accuracy by group:")
print(exp_h.groupby("group")["correct"].agg(["mean","count"]).round(3).to_string())

# ── load LLM data ──────────────────────────────────────────────────────────────
llm = pd.read_csv(LLM_CSV, encoding="utf-8-sig")
llm["correct"] = llm["correct"].map({"True": True, "False": False, True: True, False: False})
llm = llm.merge(key_df[["po_id", "experiment_block"]], on="po_id", how="left")
llm_t5 = llm[(llm["temperature"] == 0.5) & (llm["persona"] == "student")]

print(f"\nLLM (T=0.5, student) rows: {len(llm_t5)}")
print("LLM accuracy by group:")
print(llm_t5.groupby("group")["correct"].agg(["mean","count"]).round(3).to_string())

# ── load Stage 4 AI verdicts ───────────────────────────────────────────────────
with open(G2V) as f:
    g2v = json.load(f)
ai_judgment = {pid: v["judgment"] for pid, v in g2v.items()}
ai_correct  = {pid: (("anomaly" if j == "suspicious" else "normal") == truth_map.get(pid, "?"))
               for pid, j in ai_judgment.items()}
ai_wrong_pids = {pid for pid, ok in ai_correct.items() if not ok}

print(f"\nAI: {sum(ai_correct.values())}/{len(ai_correct)} correct")
print(f"AI wrong orders ({len(ai_wrong_pids)}): {sorted(ai_wrong_pids)}")

# ── helpers ────────────────────────────────────────────────────────────────────
def cliffs_d(a, b):
    a, b = list(a), list(b)
    if not a or not b:
        return float("nan")
    gt = sum(ai > bi for ai in a for bi in b)
    lt = sum(ai < bi for ai in a for bi in b)
    return (gt - lt) / (len(a) * len(b))

def boot_ci(a, b, B=5000, seed=42):
    rng = np.random.default_rng(seed)
    a, b = np.array(a, float), np.array(b, float)
    if len(a) < 2 or len(b) < 2:
        return float("nan"), float("nan")
    boots = [cliffs_d(rng.choice(a, len(a), replace=True),
                      rng.choice(b, len(b), replace=True)) for _ in range(B)]
    lo, hi = np.percentile(boots, [2.5, 97.5])
    return lo, hi

def hedges_g(a, b):
    a, b = np.array(a, float), np.array(b, float)
    na, nb = len(a), len(b)
    if na < 2 or nb < 2:
        return float("nan")
    ps = np.sqrt(((na - 1) * a.std(ddof=1) ** 2 + (nb - 1) * b.std(ddof=1) ** 2) / (na + nb - 2))
    if ps == 0:
        return 0.0
    cf = 1 - 3 / (4 * (na + nb - 2) - 1)
    return (a.mean() - b.mean()) / ps * cf

def h_group_stats(group):
    sub = exp_h[exp_h["group"] == group]
    tp = ((sub["pred"] == "anomaly") & (sub["truth"] == "anomaly")).sum()
    fn = ((sub["pred"] == "normal")  & (sub["truth"] == "anomaly")).sum()
    fp = ((sub["pred"] == "anomaly") & (sub["truth"] == "normal")).sum()
    tn = ((sub["pred"] == "normal")  & (sub["truth"] == "normal")).sum()
    n  = len(sub)
    return {
        "N_participants": sub["participant_id"].nunique(),
        "N_decisions":    n,
        "Accuracy":       sub["correct"].mean() if n > 0 else float("nan"),
        "Recall":         tp / (tp + fn) if (tp + fn) > 0 else float("nan"),
        "Specificity":    tn / (tn + fp) if (tn + fp) > 0 else float("nan"),
        "Sus_rate":       sub["is_sus"].mean() if n > 0 else float("nan"),
        "TP": tp, "FN": fn, "FP": fp, "TN": tn,
    }

def llm_group_stats(group):
    sub = llm_t5[llm_t5["group"] == group]
    if len(sub) == 0:
        return {"Accuracy": float("nan"), "Recall": float("nan"),
                "Specificity": float("nan"), "Sus_rate": float("nan"),
                "TP": 0, "FN": 0, "FP": 0, "TN": 0}
    tp = int(((sub["judgment"] == "suspicious") & (sub["truth"] == "anomaly")).sum())
    fn = int(((sub["judgment"] == "normal")     & (sub["truth"] == "anomaly")).sum())
    fp = int(((sub["judgment"] == "suspicious") & (sub["truth"] == "normal")).sum())
    tn = int(((sub["judgment"] == "normal")     & (sub["truth"] == "normal")).sum())
    return {
        "Accuracy":    sub["correct"].mean(),
        "Recall":      tp / (tp + fn) if (tp + fn) > 0 else float("nan"),
        "Specificity": tn / (tn + fp) if (tn + fp) > 0 else float("nan"),
        "Sus_rate":    (sub["judgment"] == "suspicious").mean(),
        "TP": tp, "FN": fn, "FP": fp, "TN": tn,
        "N_decisions": len(sub),
    }

# ── TABLE 1 — AI Baseline ──────────────────────────────────────────────────────
n_ai   = len(ai_correct)
sus_n  = sum(1 for j in ai_judgment.values() if j == "suspicious")
tp_ai  = sum(1 for pid, ok in ai_correct.items() if ok and truth_map[pid] == "anomaly")
tn_ai  = sum(1 for pid, ok in ai_correct.items() if ok and truth_map[pid] == "normal")
fp_ai  = sum(1 for pid, ok in ai_correct.items() if not ok and truth_map[pid] == "normal")
fn_ai  = sum(1 for pid, ok in ai_correct.items() if not ok and truth_map[pid] == "anomaly")

t1 = pd.DataFrame([
    ["Total orders evaluated",                         "32"],
    ["Accuracy",                                       f"{sum(ai_correct.values())}/{n_ai}  ({sum(ai_correct.values())/n_ai:.1%})"],
    ["Suspicious rate (proportion flagged)",           f"{sus_n}/{n_ai}  ({sus_n/n_ai:.1%})"],
    ["True Positives (TP) — correctly flagged anomalies", str(tp_ai)],
    ["True Negatives (TN) — correctly passed normals",    str(tn_ai)],
    ["False Positives (FP) — normal orders flagged",      str(fp_ai)],
    ["False Negatives (FN) — anomalies missed",           str(fn_ai)],
    ["Recall (sensitivity)",  f"{tp_ai/(tp_ai+fn_ai):.3f}" if (tp_ai + fn_ai) else "—"],
    ["Specificity",           f"{tn_ai/(tn_ai+fp_ai):.3f}" if (tn_ai + fp_ai) else "—"],
    ["Note on errors", f"AI wrong on {len(ai_wrong_pids)} orders  ({fp_ai} FP, {fn_ai} FN)"],
], columns=["Metric", "Value"])

# ── TABLE 2 — Accuracy Overview (Human PRIMARY, LLM reference) ─────────────────
print("\n── TABLE 2 ──")
t2_rows = []
for g, label in [("G1", "No AI (G1)"), ("G2", "AI Conclusion (G2)"), ("G3", "AI Evidence (G3)")]:
    hs = h_group_stats(g)
    ls = llm_group_stats(g)
    print(f"  {label}: Human acc={hs['Accuracy']:.3f}  LLM acc={ls['Accuracy']:.3f}")
    t2_rows.append({
        "Condition":               label,
        # ── HUMAN (primary) ──
        "Human N-participants":    hs["N_participants"],
        "Human N-decisions":       hs["N_decisions"],
        "Human Accuracy":          f"{hs['Accuracy']:.3f}" if not np.isnan(hs["Accuracy"]) else "—",
        "Human Recall":            f"{hs['Recall']:.3f}"   if not np.isnan(hs["Recall"])   else "—",
        "Human Specificity":       f"{hs['Specificity']:.3f}" if not np.isnan(hs["Specificity"]) else "—",
        "Human Sus. Rate":         f"{hs['Sus_rate']:.3f}" if not np.isnan(hs["Sus_rate"]) else "—",
        "Human TP|FN|FP|TN":       f"{hs['TP']}|{hs['FN']}|{hs['FP']}|{hs['TN']}",
        # ── LLM (reference) ──
        "LLM N-decisions":         ls.get("N_decisions", "—"),
        "LLM Accuracy (T=0.5)":    f"{ls['Accuracy']:.3f}" if not np.isnan(ls["Accuracy"]) else "—",
        "LLM Recall":              f"{ls['Recall']:.3f}"   if not np.isnan(ls["Recall"])   else "—",
        "LLM Specificity":         f"{ls['Specificity']:.3f}" if not np.isnan(ls["Specificity"]) else "—",
        "LLM Sus. Rate":           f"{ls['Sus_rate']:.3f}" if not np.isnan(ls["Sus_rate"]) else "—",
        "LLM TP|FN|FP|TN":         f"{ls['TP']}|{ls['FN']}|{ls['FP']}|{ls['TN']}",
    })

t2 = pd.DataFrame(t2_rows)

# ── TABLE 3 — Effect Sizes (Human PRIMARY, LLM reference) ─────────────────────
# Human: per-participant accuracy (N=4 per group)
h_accs = {g: exp_h[exp_h["group"] == g].groupby("participant_id")["correct"].mean().values
           for g in ["G1", "G2", "G3"]}
# LLM: per-question accuracy (N=32 per group, mean over 10 runs)
l_accs = {g: llm_t5[llm_t5["group"] == g].groupby("po_id")["correct"].mean().values
           for g in ["G1", "G2", "G3"]}

print("\n── TABLE 3 ──")
t3_rows = []
for ga, gb in [("G3", "G1"), ("G3", "G2"), ("G2", "G1")]:
    hd       = cliffs_d(h_accs[ga], h_accs[gb])
    hlo, hhi = boot_ci(h_accs[ga], h_accs[gb])
    hg       = hedges_g(h_accs[ga], h_accs[gb])
    ld       = cliffs_d(l_accs[ga], l_accs[gb])
    llo, lhi = boot_ci(l_accs[ga], l_accs[gb])
    lg       = hedges_g(l_accs[ga], l_accs[gb])
    print(f"  {ga} vs {gb}: Human d={hd:+.3f} [{hlo:+.3f},{hhi:+.3f}]  LLM d={ld:+.3f} [{llo:+.3f},{lhi:+.3f}]")
    t3_rows.append({
        "Comparison":             f"{ga} vs {gb}",
        # ── HUMAN (primary) ──
        "Human Cliff's d":        f"{hd:+.3f}" if not np.isnan(hd) else "—",
        "Human 95% CI":           f"[{hlo:+.3f}, {hhi:+.3f}]" if not np.isnan(hlo) else "—",
        "Human Hedges' g":        f"{hg:+.3f}" if not np.isnan(hg) else "—",
        "Human N (per group)":    f"{len(h_accs[ga])} vs {len(h_accs[gb])}",
        # ── LLM (reference) ──
        "LLM Cliff's d (T=0.5)":  f"{ld:+.3f}" if not np.isnan(ld) else "—",
        "LLM 95% CI":             f"[{llo:+.3f}, {lhi:+.3f}]" if not np.isnan(llo) else "—",
        "LLM Hedges' g":          f"{lg:+.3f}" if not np.isnan(lg) else "—",
        "LLM N (per group)":      f"{len(l_accs[ga])} vs {len(l_accs[gb])}",
    })

t3 = pd.DataFrame(t3_rows)

# ── TABLE 4 — Signal Type (Human PRIMARY, LLM reference) ─────────────────────
BLOCK_LABELS = {
    "B":   "Numerical-dominant (B)",
    "C2a": "Text-dominant (C2a)",
    "C2b": "Mixed signal (C2b)",
    "A":   "Normal routine (A)",
    "C1":  "Normal hi-Mahal (C1)",
}

print("\n── TABLE 4 ──")
t4_rows = []
for blk, label in BLOCK_LABELS.items():
    n_qs = (key_df["experiment_block"] == blk).sum()
    row  = {"Signal Type": label, "N Orders": n_qs}
    for g in ["G1", "G2", "G3"]:
        sub = exp_h[(exp_h["group"] == g) & (exp_h["experiment_block"] == blk)]
        val = f"{sub['correct'].mean():.3f}" if len(sub) > 0 else "—"
        row[f"Human {g}"] = val
    for g in ["G1", "G2", "G3"]:
        sub = llm_t5[(llm_t5["group"] == g) & (llm_t5["experiment_block"] == blk)]
        if len(sub) > 0:
            q_acc = sub.groupby("po_id")["correct"].mean().values
            row[f"LLM {g}"] = f"{q_acc.mean():.3f}"
        else:
            row[f"LLM {g}"] = "—"
    t4_rows.append(row)
    print(f"  {label}: H-G1={row['Human G1']} H-G2={row['Human G2']} H-G3={row['Human G3']} "
          f"| L-G1={row['LLM G1']} L-G2={row['LLM G2']} L-G3={row['LLM G3']}")

t4 = pd.DataFrame(t4_rows)

# ── TABLE 5 — AOR (Human PRIMARY, LLM reference) ─────────────────────────────
wrong_h = exp_h[exp_h["po_id"].isin(ai_wrong_pids)]
wrong_l = llm_t5[llm_t5["po_id"].isin(ai_wrong_pids)]

print("\n── TABLE 5 ──")
t5_rows = []
for g, label in [("G1", "No AI (G1)"), ("G2", "AI Conclusion (G2)"), ("G3", "AI Evidence (G3)")]:
    hw   = wrong_h[wrong_h["group"] == g]
    haor = hw["correct"].mean() if len(hw) > 0 else float("nan")
    lw   = wrong_l[wrong_l["group"] == g]
    laor = lw["correct"].mean() if len(lw) > 0 else float("nan")
    print(f"  {label}: Human AOR={haor:.3f} ({int(hw['correct'].sum())}/{len(hw)})  "
          f"LLM AOR={laor:.3f} ({int(lw['correct'].sum())}/{len(lw)})")
    t5_rows.append({
        "Condition":                   label,
        # ── HUMAN (primary) ──
        "Human AOR":                   f"{haor:.3f}" if not np.isnan(haor) else "—",
        "Human correct / AI-wrong orders": f"{int(hw['correct'].sum())}/{len(hw)}",
        # ── LLM (reference) ──
        "LLM AOR (T=0.5)":             f"{laor:.3f}" if not np.isnan(laor) else "—",
        "LLM correct / AI-wrong orders": f"{int(lw['correct'].sum())}/{len(lw)}",
    })

t5 = pd.DataFrame(t5_rows)

# ── TABLE 6 — Per-order Override (Human PRIMARY, LLM reference) ───────────────
print("\n── TABLE 6 ──")
t6_rows = []
for pid in sorted(ai_wrong_pids):
    blk = key_df.loc[key_df["po_id"] == pid, "experiment_block"].values[0]
    inj = key_df.loc[key_df["po_id"] == pid, "injection_plan"].values[0]
    ai_j = ai_judgment.get(pid, "?")
    tr   = truth_map.get(pid, "?")
    row  = {"PO ID": pid, "Block": blk, "AI Verdict": ai_j,
            "Ground Truth": tr, "Anomaly Type": inj}
    for g in ["G1", "G2", "G3"]:
        sub = exp_h[(exp_h["po_id"] == pid) & (exp_h["group"] == g)]
        row[f"Human {g}"] = f"{sub['correct'].mean():.2f}" if len(sub) > 0 else "—"
    for g in ["G1", "G2", "G3"]:
        sub = wrong_l[(wrong_l["po_id"] == pid) & (wrong_l["group"] == g)]
        row[f"LLM {g}"] = f"{sub['correct'].mean():.2f}" if len(sub) > 0 else "—"
    t6_rows.append(row)
    print(f"  {pid}: H-G1={row['Human G1']} H-G2={row['Human G2']} H-G3={row['Human G3']} "
          f"| L-G1={row['LLM G1']} L-G2={row['LLM G2']} L-G3={row['LLM G3']}")

t6 = pd.DataFrame(t6_rows)

# ── WRITE XLSX ─────────────────────────────────────────────────────────────────
HUMAN_COLS = {
    "T2": ["Human N-participants", "Human N-decisions", "Human Accuracy",
           "Human Recall", "Human Specificity", "Human Sus. Rate", "Human TP|FN|FP|TN"],
    "T3": ["Human Cliff's d", "Human 95% CI", "Human Hedges' g", "Human N (per group)"],
    "T4": ["Human G1", "Human G2", "Human G3"],
    "T5": ["Human AOR", "Human correct / AI-wrong orders"],
    "T6": ["Human G1", "Human G2", "Human G3"],
}

def style_sheet(ws, title, note_key, notes=None):
    """
    Layout expectation (no insert_rows used here):
      Row 1  : title        ← written by to_excel startrow=0? No —
      Row 1  : title        ← ws.cell(1,1) set below
      Row 2  : empty gap
      Row 3  : column headers  ← written by to_excel(startrow=2, index=False)
      Row 4+ : data rows
      Last+2 : footnote
    """
    thin  = Side(style="thin")
    thick = Side(style="medium")

    # ── Row 1: title (overwrite whatever pandas put there) ──────────────────
    ws.cell(1, 1).value     = title
    ws.cell(1, 1).font      = Font(bold=True, size=12)
    ws.cell(1, 1).alignment = Alignment(horizontal="left")
    # Row 2 stays empty as visual gap

    # ── Identify human-primary columns (row 3 = header row) ─────────────────
    human_col_names  = HUMAN_COLS.get(note_key, [])
    header_row_vals  = [c.value for c in ws[3]]
    human_col_idx    = {i + 1 for i, h in enumerate(header_row_vals) if h in human_col_names}

    # ── Style header row (row 3) ─────────────────────────────────────────────
    for col_idx, cell in enumerate(ws[3], 1):
        is_human = col_idx in human_col_idx
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor="1F4E79" if is_human else "4D4D4D")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = Border(top=thick, bottom=thick, left=thin, right=thin)

    # ── Style data rows (row 4 onwards) ─────────────────────────────────────
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        alt = (row[0].row % 2 == 0)
        for cell in row:
            is_human = cell.column in human_col_idx
            if is_human:
                cell.fill = PatternFill("solid", fgColor="C5D8F6" if alt else "E8F0FE")
                cell.font = Font(bold=True, size=10)
            else:
                cell.fill = PatternFill("solid", fgColor="EBEBEB" if alt else "F5F5F5")
                cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Auto column widths ───────────────────────────────────────────────────
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(max_len + 3, 12), 40)

    # ── Footnote ─────────────────────────────────────────────────────────────
    if notes:
        row_n = ws.max_row + 2
        ws.cell(row_n, 1).value     = notes
        ws.cell(row_n, 1).font      = Font(italic=True, size=9)
        ws.cell(row_n, 1).alignment = Alignment(wrap_text=True)
        if ws.max_column > 1:
            ws.merge_cells(start_row=row_n, start_column=1,
                           end_row=row_n, end_column=ws.max_column)

NOTES = {
    "T1": ("Note. Stage 4 AI = Qwen3-8B (T = 0, deterministic). N = 32 experimental orders. "
           "AI receives calibration context (company norms) but NOT Section E deviation text "
           "and NOT the 8-pattern fraud taxonomy."),
    "T2": ("Note. Human data = PRIMARY (shaded blue). N = 4 participants per group, each judging "
           "32 orders (128 decisions per group). Accuracy = correct decisions ÷ total decisions. "
           "LLM data = computational reference (shaded grey): T = 0.5, 10 stochastic runs, "
           "accuracy averaged per question then across 32 questions per group. "
           "Sus. Rate = proportion of responses labelled Suspicious."),
    "T3": ("Note. Human data = PRIMARY (shaded blue). Per-participant accuracy used for humans (N = 4 "
           "per group); per-question mean accuracy used for LLM (N = 32 questions per group). "
           "95% CI via bootstrap (B = 5,000). Magnitude thresholds: |d| < 0.11 negligible, "
           "0.11–0.28 small, 0.28–0.43 medium, > 0.43 large. Wide CIs reflect small N; "
           "results are directional, not confirmatory."),
    "T4": ("Note. Human data = PRIMARY (shaded blue). Values are mean accuracy per condition "
           "within each signal-type block. Block definitions: B = numerical-dominant anomalies; "
           "C2a = text-dominant anomalies; C2b = mixed-signal anomalies; "
           "A = routine normal orders; C1 = statistically ambiguous normal orders."),
    "T5": ("Note. Human data = PRIMARY (shaded blue). AI Override Rate (AOR) = proportion of "
           "correct answers given on orders that Stage 4 AI misclassified. "
           "Stage 4 AI was wrong on 9 orders (8 FN + 1 FP). "
           "Human denominator: 4 participants × 9 orders = 36 per group. "
           "LLM denominator: 10 runs × 9 orders = 90 per group."),
    "T6": ("Note. Human data = PRIMARY (shaded blue). Per-order correct rate = proportion of "
           "participants (Human) or simulation runs (LLM) giving correct answer for that order. "
           "All AI errors are FN (AI said Normal for anomaly) except PO-2024-0353 which is FP. "
           "Higher values = more successful override of AI error."),
}

OUT.parent.mkdir(parents=True, exist_ok=True)
SHEET_INFO = [
    ("T1 AI Baseline",       t1, "Table 1. Stage 4 AI Baseline Performance",                           "T1"),
    ("T2 Accuracy Overview", t2, "Table 2. Accuracy by Format Condition — Human (Primary) & LLM",      "T2"),
    ("T3 Effect Sizes",      t3, "Table 3. Pairwise Effect Sizes — Human (Primary) & LLM",             "T3"),
    ("T4 Signal Type",       t4, "Table 4. Accuracy by Signal Type — Human (Primary) & LLM (RQ2)",     "T4"),
    ("T5 AOR",               t5, "Table 5. AI Override Rate by Format Condition — Human (Primary) & LLM (RQ3)", "T5"),
    ("T6 Per-order AOR",     t6, "Table 6. Per-Order Override Performance — Human (Primary) & LLM (RQ3)", "T6"),
]

with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
    for sheet_name, df, title, note_key in SHEET_INFO:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        ws = writer.sheets[sheet_name]
        style_sheet(ws, title, note_key, NOTES[note_key])

print(f"\n✓ Done → {OUT}")
print("\nSheet summary:")
for name, df, title, _ in SHEET_INFO:
    print(f"  {name}: {df.shape[0]} rows × {df.shape[1]} cols")
